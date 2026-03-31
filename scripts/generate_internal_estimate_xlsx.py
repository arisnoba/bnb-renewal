from __future__ import annotations

import sys
from pathlib import Path


THIRD_PARTY = Path("/tmp/codex-openpyxl")
if THIRD_PARTY.exists():
    sys.path.insert(0, str(THIRD_PARTY))

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_PATH = Path("plan/BNB-리뉴얼-내부견적서.xlsx")

ROLE_RATES = [
    ("UX 디자이너", 350000, 450000, 600000),
    ("UI 디자이너", 300000, 400000, 550000),
    ("프론트엔드 개발", 400000, 500000, 650000),
    ("백엔드/CMS 개발", 400000, 500000, 650000),
    ("PM", 300000, 400000, 500000),
    ("QA", 250000, 300000, 400000),
]

ITEMS = [
    ("a) 마스터 템플릿", "디자인", "디자인 시스템 (색상/타이포/간격/그리드)", "UX 디자이너", 5, 1500000, 2000000, 3000000, ""),
    ("a) 마스터 템플릿", "디자인", "메인페이지 디자인 (PC+Mobile)", "UX 디자이너", 5, 1500000, 2250000, 3000000, ""),
    ("a) 마스터 템플릿", "디자인", "공통 16종 서브페이지 (static 8, gallery 4, board 2, form 2)", "UI 디자이너", 16, 4800000, 6400000, 8800000, ""),
    ("a) 마스터 템플릿", "디자인", "공통 컴포넌트 5종 (Gallery, Board, Form, Calendar, Static)", "UI 디자이너", 5, 1500000, 2000000, 2750000, ""),
    ("a) 마스터 템플릿", "디자인", "반응형 Mobile 보완 작업", "UI 디자이너", 5, 1500000, 2000000, 2750000, ""),
    ("a) 마스터 템플릿", "디자인", "프로토타이핑/인터랙션 정의", "UX 디자이너", 3, 900000, 1350000, 1800000, ""),
    ("a) 마스터 템플릿", "프론트엔드 개발", "Next.js + Monorepo 프로젝트 셋업", "프론트엔드 개발", 3, 1200000, 1500000, 1950000, ""),
    ("a) 마스터 템플릿", "프론트엔드 개발", "공통 레이아웃 (Header/GNB/Footer/LNB)", "프론트엔드 개발", 5, 2000000, 2500000, 3250000, ""),
    ("a) 마스터 템플릿", "프론트엔드 개발", "공통 16종 페이지 개발", "프론트엔드 개발", 16, 6400000, 8000000, 10400000, ""),
    ("a) 마스터 템플릿", "프론트엔드 개발", "5종 페이지 타입 컴포넌트 라이브러리", "프론트엔드 개발", 8, 3200000, 4000000, 5200000, ""),
    ("a) 마스터 템플릿", "프론트엔드 개발", "하이틴센터 고유 5종 (특강, 다이렉트캐스팅, 섭외뉴스 등)", "프론트엔드 개발", 5, 2000000, 2500000, 3250000, ""),
    ("a) 마스터 템플릿", "프론트엔드 개발", "반응형 처리 + 크로스브라우저", "프론트엔드 개발", 4, 1600000, 2000000, 2600000, ""),
    ("a-1) 브랜딩 리디자인", "브랜드 전략", "5개 센터 BI 방향성 수립 + 리서치", "UX 디자이너", 3, 900000, 1350000, 1800000, ""),
    ("a-1) 브랜딩 리디자인", "브랜드 전략", "로고/심볼/컬러 시스템 재정립 (5개)", "UX 디자이너", 10, 3000000, 4500000, 6000000, ""),
    ("a-1) 브랜딩 리디자인", "브랜드 전략", "톤앤매너 가이드 (사진/일러스트/아이콘 스타일)", "UI 디자이너", 5, 1500000, 2000000, 2750000, ""),
    ("a-1) 브랜딩 리디자인", "브랜드 전략", "브랜드 가이드라인 문서 제작", "UX 디자이너", 3, 900000, 1350000, 1650000, ""),
    ("b) 브랜드 커스터마이징", "아트센터", "브랜드 디자인 적용 (색상/이미지/로고 교체)", "UI 디자이너", "", 600000, 800000, 1100000, ""),
    ("b) 브랜드 커스터마이징", "아트센터", "고유 5종 디자인 (센터소개, 매니저소개, 출신아티스트, 촬영후기, 오디션폼)", "UI 디자이너", "", 1500000, 2000000, 2750000, ""),
    ("b) 브랜드 커스터마이징", "아트센터", "고유 5종 개발", "프론트엔드 개발", "", 2000000, 2500000, 3250000, ""),
    ("b) 브랜드 커스터마이징", "입시센터", "브랜드 디자인 적용", "UI 디자이너", "", 600000, 800000, 1100000, ""),
    ("b) 브랜드 커스터마이징", "입시센터", "고유 13종 디자인 (커리큘럼 5, 합격현황 2, 후기/영상, 장학, 연혁 등)", "UI 디자이너", "", 3900000, 5200000, 7150000, ""),
    ("b) 브랜드 커스터마이징", "입시센터", "고유 13종 개발", "프론트엔드 개발", "", 5200000, 6500000, 8450000, ""),
    ("b) 브랜드 커스터마이징", "키즈센터", "브랜드 디자인 적용", "UI 디자이너", "", 600000, 800000, 1100000, ""),
    ("b) 브랜드 커스터마이징", "키즈센터", "고유 5종 디자인 (교육과정 3, 추가 캐스팅보드 2)", "UI 디자이너", "", 1500000, 2000000, 2750000, ""),
    ("b) 브랜드 커스터마이징", "키즈센터", "고유 5종 개발", "프론트엔드 개발", "", 2000000, 2500000, 3250000, ""),
    ("c) 공통 기능", "CMS/플랫폼", "Payload CMS 셋업 + 콘텐츠 모델링 (5개 브랜드)", "백엔드/CMS 개발", 10, 4000000, 5000000, 6500000, ""),
    ("c) 공통 기능", "CMS/플랫폼", "CMS 어드민 UI 커스터마이징 + 권한 관리", "백엔드/CMS 개발", 5, 2000000, 2500000, 3250000, ""),
    ("c) 공통 기능", "CMS/플랫폼", "통합 상담 폼 (5개 센터 선택 + API + 알림)", "백엔드/CMS 개발", 5, 2000000, 2500000, 3250000, ""),
    ("c) 공통 기능", "CMS/플랫폼", "통합 관리자 대시보드 (센터별 상담 조회/관리)", "백엔드/CMS 개발", 6, 2400000, 3000000, 3900000, ""),
    ("c) 공통 기능", "CMS/플랫폼", "그누보드 데이터 마이그레이션 스크립트", "백엔드/CMS 개발", 5, 2000000, 2500000, 3250000, ""),
    ("c) 공통 기능", "CMS/플랫폼", "공통 API 레이어 + 타입 공유 (Monorepo)", "백엔드/CMS 개발", 4, 1600000, 2000000, 2600000, ""),
    ("d) SEO/GEO 인프라", "SEO/GEO", "클린 URL 설계 + 301 리다이렉트 맵 (126페이지 전체)", "프론트엔드 개발", "", 1000000, 1500000, 2000000, ""),
    ("d) SEO/GEO 인프라", "SEO/GEO", "Schema.org 구조화 데이터 (5개 사이트 × 5종 스키마)", "프론트엔드 개발", "", 2000000, 3000000, 4000000, ""),
    ("d) SEO/GEO 인프라", "SEO/GEO", "페이지별 동적 메타 (title/description/OG/canonical)", "프론트엔드 개발", "", 1000000, 1500000, 2000000, ""),
    ("d) SEO/GEO 인프라", "SEO/GEO", "sitemap.xml 자동 생성 + robots.txt (5개 사이트)", "프론트엔드 개발", "", 500000, 800000, 1000000, ""),
    ("d) SEO/GEO 인프라", "SEO/GEO", "llms.txt 작성 (5개 사이트)", "프론트엔드 개발", "", 500000, 800000, 1000000, ""),
    ("d) SEO/GEO 인프라", "SEO/GEO", "Core Web Vitals 최적화 (LCP/CLS/INP)", "프론트엔드 개발", "", 1000000, 1500000, 2000000, ""),
    ("d) SEO/GEO 인프라", "SEO/GEO", "Google Search Console + 사이트맵 제출 (5개)", "PM", "", 500000, 800000, 1000000, ""),
    ("e) 에비뉴센터 별도", "에비뉴센터", "SSL 인증서 갱신 + 서버 정비", "백엔드/CMS 개발", "", 300000, 500000, 500000, ""),
    ("e) 에비뉴센터 별도", "에비뉴센터", "원페이지 랜딩 디자인 (PC+Mobile, 4섹션)", "UI 디자이너", "", 1500000, 2500000, 3500000, ""),
    ("e) 에비뉴센터 별도", "에비뉴센터", "서브페이지 4종 디자인 (FAQ, 강사, 수강안내, 상담)", "UI 디자이너", "", 1200000, 1600000, 2200000, ""),
    ("e) 에비뉴센터 별도", "에비뉴센터", "프론트엔드 개발 (랜딩 + 서브 4종)", "프론트엔드 개발", "", 2000000, 3000000, 4000000, ""),
    ("e) 에비뉴센터 별도", "에비뉴센터", "SEO/GEO 기본 셋업", "프론트엔드 개발", "", 1000000, 1500000, 2000000, ""),
    ("f) 기타", "기타", "프로젝트 관리 (PM, 회의, 보고)", "PM", "", 1500000, 2500000, 3500000, ""),
    ("f) 기타", "기타", "QA 및 테스팅 (5개 사이트 × 크로스브라우저)", "QA", "", 2500000, 3500000, 4500000, ""),
    ("f) 기타", "기타", "호스팅/인프라 셋업 (Vercel 또는 자체 서버)", "백엔드/CMS 개발", "", 500000, 1000000, 1500000, ""),
    ("f) 기타", "기타", "호스팅 연간 운영비", "", "", 1200000, 2400000, 3600000, "직군 단가 비연동"),
    ("f) 기타", "기타", "콘텐츠 입력 지원 + 데이터 정리", "PM", "", 1000000, 2000000, 3000000, ""),
    ("f) 기타", "기타", "운영 가이드 문서 + 클라이언트 교육", "PM", "", 500000, 1000000, 1500000, ""),
]

SUMMARY_GROUPS = [
    "a) 마스터 템플릿",
    "a-1) 브랜딩 리디자인",
    "b) 브랜드 커스터마이징",
    "c) 공통 기능",
    "d) SEO/GEO 인프라",
    "e) 에비뉴센터 별도",
    "f) 기타",
]


def apply_base_style(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


def set_common_styles():
    thin = Side(style="thin", color="D9D9D9")
    return {
        "header_fill": PatternFill("solid", fgColor="1F4E78"),
        "sub_fill": PatternFill("solid", fgColor="D9EAF7"),
        "input_fill": PatternFill("solid", fgColor="FFF2CC"),
        "formula_fill": PatternFill("solid", fgColor="E2F0D9"),
        "title_fill": PatternFill("solid", fgColor="DDEBF7"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "header_font": Font(color="FFFFFF", bold=True),
        "bold_font": Font(bold=True),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
    }


def add_settings_sheet(wb, styles):
    ws = wb.active
    ws.title = "단가설정"
    apply_base_style(ws)

    ws["A1"] = "BNB 리뉴얼 내부 견적서 입력값"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = styles["title_fill"]

    ws["A3"] = "입력 안내"
    ws["A4"] = "1. 노란색 셀만 수정하면 됩니다."
    ws["A5"] = "2. 직군 단가를 수정하면 견적입력 시트의 연동 항목 금액이 비례 반영됩니다."
    ws["A6"] = "3. 견적입력 시트에서 포함(Y/N), 기준금액, 수동금액을 바꾸면 요약이 자동 갱신됩니다."

    ws["A8"] = "VAT"
    ws["B8"] = 0.1
    ws["B8"].number_format = "0%"
    ws["B8"].fill = styles["input_fill"]

    ws["A10"] = "직군"
    ws["B10"] = "LOW"
    ws["C10"] = "MID"
    ws["D10"] = "HIGH"

    for cell in ws["A10:D10"][0]:
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    for row_idx, (role, low, mid, high) in enumerate(ROLE_RATES, start=11):
        ws[f"A{row_idx}"] = role
        ws[f"B{row_idx}"] = low
        ws[f"C{row_idx}"] = mid
        ws[f"D{row_idx}"] = high
        for col in "ABCD":
            ws[f"{col}{row_idx}"].border = styles["border"]
        for col in "BCD":
            ws[f"{col}{row_idx}"].fill = styles["input_fill"]
            ws[f"{col}{row_idx}"].number_format = '#,##0"원"'

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    return ws


def build_rate_lookup_formula(level_col: str, row_num: int) -> str:
    return (
        f'IF($E{row_num}="","",IFERROR(INDEX(단가설정!${level_col}$11:${level_col}$16,'
        f'MATCH($E{row_num},단가설정!$A$11:$A$16,0)),""))'
    )


def add_input_sheet(wb, styles):
    ws = wb.create_sheet("견적입력")
    apply_base_style(ws)

    headers = [
        "포함",
        "대분류",
        "세부구분",
        "항목",
        "연동직군",
        "공수(일)",
        "LOW 기준금액",
        "MID 기준금액",
        "HIGH 기준금액",
        "LOW 수동금액",
        "MID 수동금액",
        "HIGH 수동금액",
        "LOW 반영금액",
        "MID 반영금액",
        "HIGH 반영금액",
        "비고",
        "기준 LOW단가",
        "기준 MID단가",
        "기준 HIGH단가",
    ]

    ws["A1"] = "BNB 리뉴얼 내부 견적 입력"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = styles["title_fill"]
    ws["A2"] = "노란색 셀은 입력용, 초록색 셀은 자동 계산입니다. 포함 열에서 Y를 선택한 항목만 합계에 반영됩니다."

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    yn_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    ws.add_data_validation(yn_validation)

    currency_cols = ["G", "H", "I", "J", "K", "L", "M", "N", "O", "Q", "R", "S"]
    widths = {
        "A": 8,
        "B": 22,
        "C": 16,
        "D": 56,
        "E": 18,
        "F": 10,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14,
        "O": 14,
        "P": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    start_row = 4
    for idx, (group, subgroup, item, role, days, low, mid, high, note) in enumerate(ITEMS, start=start_row):
        ws[f"A{idx}"] = "Y"
        ws[f"B{idx}"] = group
        ws[f"C{idx}"] = subgroup
        ws[f"D{idx}"] = item
        ws[f"E{idx}"] = role
        ws[f"F{idx}"] = days
        ws[f"G{idx}"] = low
        ws[f"H{idx}"] = mid
        ws[f"I{idx}"] = high
        ws[f"P{idx}"] = note

        ws[f"Q{idx}"] = (
            f'=IF($E{idx}="","",IFERROR(INDEX(단가설정!$B$11:$B$16,'
            f'MATCH($E{idx},단가설정!$A$11:$A$16,0)),""))'
        )
        ws[f"R{idx}"] = (
            f'=IF($E{idx}="","",IFERROR(INDEX(단가설정!$C$11:$C$16,'
            f'MATCH($E{idx},단가설정!$A$11:$A$16,0)),""))'
        )
        ws[f"S{idx}"] = (
            f'=IF($E{idx}="","",IFERROR(INDEX(단가설정!$D$11:$D$16,'
            f'MATCH($E{idx},단가설정!$A$11:$A$16,0)),""))'
        )

        ws[f"M{idx}"] = (
            f'=IF($A{idx}<>"Y",0,IF($J{idx}<>"",$J{idx},IF($Q{idx}="",$G{idx},ROUND($G{idx}*('
            f'{build_rate_lookup_formula("B", idx)}/$Q{idx}),0))))'
        )
        ws[f"N{idx}"] = (
            f'=IF($A{idx}<>"Y",0,IF($K{idx}<>"",$K{idx},IF($R{idx}="",$H{idx},ROUND($H{idx}*('
            f'{build_rate_lookup_formula("C", idx)}/$R{idx}),0))))'
        )
        ws[f"O{idx}"] = (
            f'=IF($A{idx}<>"Y",0,IF($L{idx}<>"",$L{idx},IF($S{idx}="",$I{idx},ROUND($I{idx}*('
            f'{build_rate_lookup_formula("D", idx)}/$S{idx}),0))))'
        )

        yn_validation.add(ws[f"A{idx}"])

        for col in "ABCDEFGHIJKLMNOPQRS":
            ws[f"{col}{idx}"].border = styles["border"]
            ws[f"{col}{idx}"].alignment = styles["left"] if col in {"B", "C", "D", "E", "P"} else styles["center"]

        for col in ["A", "G", "H", "I", "J", "K", "L"]:
            ws[f"{col}{idx}"].fill = styles["input_fill"]
        for col in ["M", "N", "O"]:
            ws[f"{col}{idx}"].fill = styles["formula_fill"]

    for col in currency_cols:
        for row_idx in range(start_row, start_row + len(ITEMS)):
            ws[f"{col}{row_idx}"].number_format = '#,##0"원"'

    for row_idx in range(start_row, start_row + len(ITEMS)):
        ws[f"F{row_idx}"].number_format = '0.0"일"'

    ws.auto_filter.ref = f"A3:P{start_row + len(ITEMS) - 1}"
    ws.row_dimensions[2].height = 28
    ws.conditional_formatting.add(
        f"A{start_row}:O{start_row + len(ITEMS) - 1}",
        FormulaRule(formula=[f'$A{start_row}="N"'], fill=PatternFill("solid", fgColor="F2F2F2")),
    )
    ws.column_dimensions["Q"].hidden = True
    ws.column_dimensions["R"].hidden = True
    ws.column_dimensions["S"].hidden = True
    return ws


def add_summary_sheet(wb, styles):
    ws = wb.create_sheet("요약")
    apply_base_style(ws)

    ws["A1"] = "견적 요약"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = styles["title_fill"]

    ws["A3"] = "구분"
    ws["B3"] = "LOW"
    ws["C3"] = "MID"
    ws["D3"] = "HIGH"
    for cell in ws["A3:D3"][0]:
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    for row_idx, group in enumerate(SUMMARY_GROUPS, start=4):
        ws[f"A{row_idx}"] = group
        ws[f"B{row_idx}"] = f'=SUMIFS(견적입력!$M:$M,견적입력!$B:$B,$A{row_idx})'
        ws[f"C{row_idx}"] = f'=SUMIFS(견적입력!$N:$N,견적입력!$B:$B,$A{row_idx})'
        ws[f"D{row_idx}"] = f'=SUMIFS(견적입력!$O:$O,견적입력!$B:$B,$A{row_idx})'
        for col in "ABCD":
            ws[f"{col}{row_idx}"].border = styles["border"]

    total_row = 4 + len(SUMMARY_GROUPS)
    vat_row = total_row + 1
    vat_total_row = total_row + 2
    selected_row = total_row + 4

    ws[f"A{total_row}"] = "합계 (VAT 별도)"
    ws[f"B{total_row}"] = f"=SUM(B4:B{total_row - 1})"
    ws[f"C{total_row}"] = f"=SUM(C4:C{total_row - 1})"
    ws[f"D{total_row}"] = f"=SUM(D4:D{total_row - 1})"

    ws[f"A{vat_row}"] = "VAT"
    ws[f"B{vat_row}"] = f"=B{total_row}*단가설정!$B$8"
    ws[f"C{vat_row}"] = f"=C{total_row}*단가설정!$B$8"
    ws[f"D{vat_row}"] = f"=D{total_row}*단가설정!$B$8"

    ws[f"A{vat_total_row}"] = "VAT 포함"
    ws[f"B{vat_total_row}"] = f"=B{total_row}+B{vat_row}"
    ws[f"C{vat_total_row}"] = f"=C{total_row}+C{vat_row}"
    ws[f"D{vat_total_row}"] = f"=D{total_row}+D{vat_row}"

    ws[f"A{selected_row}"] = "선택 항목 수"
    ws[f"B{selected_row}"] = '=COUNTIF(견적입력!$A:$A,"Y")'
    ws[f"A{selected_row + 1}"] = "제외 항목 수"
    ws[f"B{selected_row + 1}"] = '=COUNTIF(견적입력!$A:$A,"N")'

    for row_idx in [total_row, vat_row, vat_total_row, selected_row, selected_row + 1]:
        for col in "ABCD":
            ws[f"{col}{row_idx}"].border = styles["border"]

    for row_idx in [total_row, vat_total_row]:
        for col in "ABCD":
            ws[f"{col}{row_idx}"].fill = styles["sub_fill"]
            ws[f"{col}{row_idx}"].font = styles["bold_font"]

    for col in "BCD":
        for row_idx in range(4, vat_total_row + 1):
            ws[f"{col}{row_idx}"].number_format = '#,##0"원"'

    ws["F3"] = "사용 방법"
    ws["F4"] = "1. 단가설정 시트의 LOW/MID/HIGH 단가를 수정"
    ws["F5"] = "2. 견적입력 시트에서 포함(Y/N)과 금액을 조정"
    ws["F6"] = "3. 요약 시트에서 최종 금액 확인"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["F"].width = 40
    return ws


def main():
    wb = Workbook()
    styles = set_common_styles()
    add_settings_sheet(wb, styles)
    add_input_sheet(wb, styles)
    add_summary_sheet(wb, styles)

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Created: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
